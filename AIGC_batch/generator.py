"""
Key生成核心逻辑模块
处理xlsx文件读取、Prompt模板替换、批量生成、断点续传
"""

import openpyxl
import json
import os
import glob
import time
import concurrent.futures
from typing import List, Dict, Any, Optional, Tuple
from dataclasses import dataclass, asdict
from datetime import datetime


@dataclass
class GenerationResult:
    """生成结果"""
    row_index: int
    input_data: Dict[str, str]
    result: str
    success: bool
    error: Optional[str] = None
    timestamp: float = 0
    generation_time: float = 0  # 生成耗时（秒）
    parsed_result: Dict[str, Any] = None  # 解析后的JSON结果

    def __post_init__(self):
        if self.timestamp == 0:
            self.timestamp = time.time()
        if self.parsed_result is None:
            self.parsed_result = {}


@dataclass
class CheckpointData:
    """断点数据"""
    timestamp: float
    total_rows: int
    current_index: int
    results: List[Dict[str, Any]]
    input_file: str
    prompt_template: str
    variable_mapping: Dict[str, str]

    def to_dict(self) -> Dict:
        return asdict(self)

    @classmethod
    def from_dict(cls, data: Dict) -> 'CheckpointData':
        return cls(**data)

    @staticmethod
    def from_generation_result(result: GenerationResult) -> Dict[str, Any]:
        """
        将GenerationResult转换为字典格式，用于保存到断点
        :param result: 生成结果对象
        :return: 字典格式的结果
        """
        return {
            "row_index": result.row_index,
            "input_data": result.input_data,
            "result": result.result,
            "success": result.success,
            "error": result.error,
            "timestamp": result.timestamp,
            "generation_time": result.generation_time,
            "parsed_result": result.parsed_result
        }


class KeyGenerator:
    """Key生成器"""

    def __init__(self, save_interval: int = 20):
        """
        初始化生成器
        :param save_interval: 自动保存间隔行数
        """
        self.save_interval = save_interval
        self.input_data: List[Dict[str, str]] = []
        self.headers: List[str] = []
        self.results: List[GenerationResult] = []
        self.total_rows: int = 0
        self._current_file: Optional[str] = None
        self._is_generating: bool = False
        self._last_checkpoint: Optional[str] = None

    def load_input(self, file_path: str, sheet_name: Optional[str] = None) -> Tuple[bool, str]:
        """
        加载输入文件
        :param file_path: xlsx文件路径
        :param sheet_name: 工作表名称，默认第一个
        :return: (成功, 消息)
        """
        try:
            if not os.path.exists(file_path):
                return False, f"文件不存在: {file_path}"

            wb = openpyxl.load_workbook(file_path)
            sheet = wb[sheet_name] if sheet_name else wb.active

            # 读取表头
            self.headers = [sheet.cell(1, col).value for col in range(1, sheet.max_column + 1)]

            # 读取数据
            self.input_data = []
            for row_idx in range(2, sheet.max_row + 1):
                row_data = {}
                for col_idx, header in enumerate(self.headers, 1):
                    row_data[header] = sheet.cell(row_idx, col_idx).value
                self.input_data.append(row_data)

            self.total_rows = len(self.input_data)
            self._current_file = file_path

            return True, f"成功加载 {self.total_rows} 行数据"

        except Exception as e:
            return False, f"加载文件失败: {e}"

    def get_data_preview(self, n: int = 5) -> List[Dict[str, str]]:
        """获取数据预览"""
        return self.input_data[:n]

    def render_prompt(self, template: str, variables: Dict[str, str], row_data: Dict[str, str]) -> str:
        """
        渲染Prompt模板
        :param template: Prompt模板
        :param variables: 变量映射 {模板变量: 列名}
        :param row_data: 行数据
        :return: 渲染后的Prompt
        """
        prompt = template
        for var_name, column_name in variables.items():
            # 支持 {{变量名}} 语法
            placeholder1 = f"{{{{{var_name}}}}}"
            # 也支持 {变量名} 语法
            placeholder2 = f"{{{var_name}}}"

            value = row_data.get(column_name, "")
            if value is None:
                value = ""

            prompt = prompt.replace(placeholder1, str(value))
            prompt = prompt.replace(placeholder2, str(value))

        return prompt

    def _parse_json_result(self, result_str: str) -> Dict[str, Any]:
        """
        尝试解析JSON结果
        :param result_str: 生成的结果字符串
        :return: 解析后的JSON字典，如果解析失败返回空字典
        """
        try:
            # 尝试直接解析
            return json.loads(result_str)
        except json.JSONDecodeError:
            # 尝试清理后再解析（移除可能的多余文本）
            try:
                # 查找JSON开始和结束位置
                start = result_str.find('{')
                end = result_str.rfind('}')
                if start != -1 and end != -1:
                    json_str = result_str[start:end+1]
                    return json.loads(json_str)
            except Exception:
                pass
            return {}

    def generate_single(self, api_client, row_index: int, template: str,
                       variables: Dict[str, str]) -> GenerationResult:
        """
        生成单行数据
        :param api_client: API客户端
        :param row_index: 行索引
        :param template: Prompt模板
        :param variables: 变量映射
        :return: 生成结果
        """
        row_data = self.input_data[row_index]

        start_time = time.time()  # 开始时间
        try:
            prompt = self.render_prompt(template, variables, row_data)
            result = api_client.generate(prompt)
            end_time = time.time()  # 结束时间
            
            # 解析JSON结果
            parsed_result = self._parse_json_result(result)
            
            return GenerationResult(
                row_index=row_index,
                input_data=row_data,
                result=result,
                success=True,
                generation_time=end_time - start_time,  # 计算耗时
                parsed_result=parsed_result  # 添加解析后的JSON
            )
        except Exception as e:
            end_time = time.time()  # 结束时间
            return GenerationResult(
                row_index=row_index,
                input_data=row_data,
                result="",
                success=False,
                error=str(e),
                generation_time=end_time - start_time  # 计算耗时
            )

    def preview_first_n(self, api_client, n: int = 3, template: str = "",
                       variables: Dict[str, str] = None) -> Tuple[List[GenerationResult], str]:
        """
        预览前N行
        :param api_client: API客户端
        :param n: 预览行数
        :param template: Prompt模板
        :param variables: 变量映射
        :return: (结果列表, 消息)
        """
        if variables is None:
            variables = {}

        preview_count = min(n, len(self.input_data))
        results = []

        for i in range(preview_count):
            result = self.generate_single(api_client, i, template, variables)
            results.append(result)

        return results, f"预览完成，共 {len(results)} 行"

    def start_generation(self, api_client, template: str, variables: Dict[str, str],
                        start_index: int = 0, on_progress=None, max_workers: int = 5) -> Tuple[bool, str]:
        """
        开始批量生成
        :param api_client: API客户端
        :param template: Prompt模板
        :param variables: 变量映射
        :param start_index: 起始索引（用于断点续传）
        :param on_progress: 进度回调
        :param max_workers: 最大并发数
        :return: (成功, 消息)
        """
        if self._is_generating:
            return False, "生成任务正在进行中"

        self._is_generating = True
        self.results = [None] * (len(self.input_data) - start_index)  # 预分配结果列表
        success_count = 0
        error_count = 0
        total_start_time = time.time()  # 总开始时间

        try:
            def process_item(index):
                """处理单个任务"""
                result = self.generate_single(api_client, index, template, variables)
                return index - start_index, result

            # 使用线程池并发处理
            with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
                # 提交所有任务
                future_to_index = {
                    executor.submit(process_item, i): i 
                    for i in range(start_index, len(self.input_data))
                }

                # 处理完成的任务
                for future in concurrent.futures.as_completed(future_to_index):
                    try:
                        result_index, result = future.result()
                        self.results[result_index] = result

                        if result.success:
                            success_count += 1
                        else:
                            error_count += 1

                        # 进度回调
                        current_completed = success_count + error_count
                        if on_progress:
                            on_progress(current_completed, len(self.input_data), success_count, error_count)

                        # 定期保存断点
                        if current_completed % self.save_interval == 0:
                            self.save_checkpoint(template, variables, start_index + current_completed)
                    except Exception as e:
                        error_count += 1
                        if on_progress:
                            on_progress(success_count + error_count, len(self.input_data), success_count, error_count)

            # 最终保存
            self.save_checkpoint(template, variables, len(self.input_data))

            total_end_time = time.time()  # 总结束时间
            total_time = total_end_time - total_start_time

            self._is_generating = False
            return True, f"生成完成，成功: {success_count}，失败: {error_count}，总耗时: {total_time:.2f}秒"

        except Exception as e:
            self._is_generating = False
            # 保存当前进度
            self.save_checkpoint(template, variables, start_index + len([r for r in self.results if r]))
            return False, f"生成中断: {e}，已保存当前进度"

    def save_checkpoint(self, template: str, variables: Dict[str, str], current_index: int):
        """保存断点"""
        checkpoint_dir = os.path.dirname(self._current_file) if self._current_file else "."
        checkpoint_path = os.path.join(
            checkpoint_dir,
            f".checkpoint_{int(time.time())}.json"
        )

        # 将结果转换为字典格式，确保parsed_result被正确保存
        results_dict = [CheckpointData.from_generation_result(r) for r in self.results if r]

        checkpoint = CheckpointData(
            timestamp=time.time(),
            total_rows=self.total_rows,
            current_index=current_index,
            results=results_dict,
            input_file=self._current_file or "",
            prompt_template=template,
            variable_mapping=variables
        )

        with open(checkpoint_path, 'w', encoding='utf-8') as f:
            json.dump(checkpoint.to_dict(), f, ensure_ascii=False, indent=2)

        self._last_checkpoint = checkpoint_path

        # 清理旧断点（保留最近3个）
        self._cleanup_old_checkpoints(checkpoint_dir, keep=3)

    def load_checkpoint(self, checkpoint_path: str) -> Tuple[bool, str]:
        """
        加载断点
        :param checkpoint_path: 断点文件路径
        :return: (成功, 消息)
        """
        try:
            if not os.path.exists(checkpoint_path):
                return False, f"断点文件不存在: {checkpoint_path}"

            with open(checkpoint_path, 'r', encoding='utf-8') as f:
                data = json.load(f)

            checkpoint = CheckpointData.from_dict(data)

            # 恢复输入数据
            if checkpoint.input_file and os.path.exists(checkpoint.input_file):
                self.load_input(checkpoint.input_file)

            # 恢复结果，确保parsed_result字段被正确处理
            self.results = []
            for r in checkpoint.results:
                # 兼容旧版断点数据（没有parsed_result字段）
                if 'parsed_result' not in r:
                    r['parsed_result'] = {}
                # 恢复GenerationResult对象
                self.results.append(GenerationResult(**r))

            return True, f"加载断点成功，当前进度: {checkpoint.current_index}/{checkpoint.total_rows}"

        except Exception as e:
            return False, f"加载断点失败: {e}"

    def get_latest_checkpoint(self, checkpoint_dir: Optional[str] = None) -> Optional[str]:
        """获取最新断点文件"""
        if checkpoint_dir is None:
            checkpoint_dir = os.path.dirname(self._current_file) if self._current_file else "."

        checkpoints = glob.glob(os.path.join(checkpoint_dir, ".checkpoint_*.json"))
        if not checkpoints:
            return None

        return max(checkpoints, key=os.path.getmtime)

    def _flatten_json(self, data: Dict[str, Any], prefix: str = '') -> Dict[str, Any]:
        """
        扁平化JSON结构
        :param data: 输入的JSON字典
        :param prefix: 前缀，用于处理嵌套结构
        :return: 扁平化后的JSON字典
        """
        flattened = {}
        for key, value in data.items():
            full_key = f"{prefix}.{key}" if prefix else key
            if isinstance(value, dict):
                flattened.update(self._flatten_json(value, full_key))
            else:
                flattened[full_key] = value
        return flattened

    def _cleanup_old_checkpoints(self, directory: str, keep: int = 3):
        """清理旧断点文件"""
        checkpoints = glob.glob(os.path.join(directory, ".checkpoint_*.json"))
        if len(checkpoints) <= keep:
            return

        # 按修改时间排序，删除最旧的
        checkpoints.sort(key=os.path.getmtime, reverse=True)
        for old_checkpoint in checkpoints[keep:]:
            try:
                os.remove(old_checkpoint)
            except:
                pass

    def export_result(self, output_path: str, input_path: Optional[str] = None) -> Tuple[bool, str]:
        """
        导出结果到xlsx
        :param output_path: 输出文件路径
        :param input_path: 输入文件路径（如果提供，则复制原文件结构）
        :return: (成功, 消息)
        """
        try:
            if input_path and os.path.exists(input_path):
                # 复制原文件并添加结果列
                wb = openpyxl.load_workbook(input_path)
                sheet = wb.active
                # 找到最后一列
                base_col = sheet.max_column + 1
            else:
                # 创建新文件
                wb = openpyxl.Workbook()
                sheet = wb.active
                # 写入表头
                if self.headers:
                    for col_idx, header in enumerate(self.headers, 1):
                        sheet.cell(1, col_idx, header)
                base_col = len(self.headers) + 1

            # 首先添加基础结果列
            sheet.cell(1, base_col, "召回Key")
            sheet.cell(1, base_col + 1, "状态")
            sheet.cell(1, base_col + 2, "错误信息")
            sheet.cell(1, base_col + 3, "生成耗时(秒)")
            
            # 收集所有解析后的JSON键
            json_keys = set()
            for result in self.results:
                if result and result.success and result.parsed_result:
                    flattened = self._flatten_json(result.parsed_result)
                    json_keys.update(flattened.keys())
            
            # 排序JSON键以便一致的列顺序
            json_keys = sorted(json_keys)
            
            # 添加JSON列的表头
            json_col = base_col + 4
            json_key_map = {}
            for i, key in enumerate(json_keys):
                sheet.cell(1, json_col + i, f"JSON.{key}")
                json_key_map[key] = json_col + i
            
            # 写入结果数据
            for result in self.results:
                if result is None:
                    continue
                
                row_idx = result.row_index + 2  # +2 因为有表头且索引从0开始
                
                # 写入基础结果
                sheet.cell(row_idx, base_col, result.result)
                sheet.cell(row_idx, base_col + 1, "成功" if result.success else "失败")
                sheet.cell(row_idx, base_col + 2, result.error or "")
                sheet.cell(row_idx, base_col + 3, f"{result.generation_time:.2f}" if result.generation_time > 0 else "")
                
                # 写入解析后的JSON数据
                if result.success and result.parsed_result:
                    flattened = self._flatten_json(result.parsed_result)
                    for key, value in flattened.items():
                        if key in json_key_map:
                            sheet.cell(row_idx, json_key_map[key], str(value))

            wb.save(output_path)
            
            if json_keys:
                return True, f"结果已保存至: {output_path} (含{len(json_keys)}个JSON字段)"
            else:
                return True, f"结果已保存至: {output_path}"

        except Exception as e:
            return False, f"导出失败: {e}"

    def get_progress(self) -> Dict[str, Any]:
        """获取当前进度"""
        completed_results = [r for r in self.results if r is not None]
        completed = len(completed_results)
        
        # 计算总生成耗时
        total_generation_time = sum(r.generation_time for r in completed_results if r.success)
        avg_generation_time = total_generation_time / len([r for r in completed_results if r.success]) if completed > 0 else 0
        
        return {
            "total": self.total_rows,
            "current": completed,
            "progress": round(completed / self.total_rows * 100, 2) if self.total_rows > 0 else 0,
            "success": sum(1 for r in completed_results if r.success),
            "error": sum(1 for r in completed_results if not r.success),
            "is_generating": self._is_generating,
            "total_generation_time": round(total_generation_time, 2),
            "avg_generation_time": round(avg_generation_time, 2)
        }

    def clear(self):
        """清空状态"""
        self.input_data = []
        self.headers = []
        self.results = []
        self.total_rows = 0
        self._current_file = None
        self._is_generating = False
